import json
import time
import requests
import os
import openpyxl

# 添加到excel

# as
def write_excel_xlsx_append(path, value_list=[[]]):
    # 判断文件是否存在
    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        workbook.create_sheet("rating", 0)
        summary = workbook['rating']
        title = ['评论id', '时间', 'rating', 'package', 'quality', 'username', '评论']
        summary.append(title)
        workbook.save(path)
    if len(value_list) != 0:
        workbook = openpyxl.load_workbook(path)
        for line in value_list:
            sheet = workbook['rating']
            sheet.append(line)
        workbook.save(path)  # 保存工作簿
    pass


def get_page(rating_url, index, pizza, sid, save_path):
    params = {
        'has_content': True,
        'tag_name': '全部',
        'offset': 200 * index,
        'limit': 200,
    }
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU iPhone OS 13_1_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.1 Mobile/15E148 Safari/604.1',
        'cookie': 'pizza7567632f76332f72=' + pizza + '; SID=' + sid + ';',
    }
    html = requests.get(rating_url, params=params, headers=headers)
    nextPizza = html.cookies['pizza7567632f76332f72']
    items = json.loads(html.text)
    row = []
    for item in items:
        if len(item) < 8:
            print('发生错误，重新开始爬', index)
            print(item)
            nextPizza = get_page(rating_url, index, pizza, sid, save_path)
            return nextPizza
        print(item['rateId'], item['rating_text'])

        row.append([item['rateId'], item['rated_at'], item['rating'], item['package_rating'], item['quality_rating'],
                    item['username'],
                    item['rating_text']])
    write_excel_xlsx_append(save_path, row)  # 保存文件名
    print(index, nextPizza)
    return nextPizza  #返回下一次请求所需nextPizza


if __name__ == '__main__':
    save_path = './三米1_2.xlsx'  #保存路径
    sid = 'XQAAAAB5plEJ7AAG3QA1ZTRmNTY2MzgzOY0Zv1P_hQ5TPuYE-QYg_Ofh'  # 登录之后从cookie中获取
    pizza = 'QUHfPno15MSPNFMYfJmlrD3t0YcPSlbrJ8EVNnCn-l5cMIJOVnYCqBmveJvVyyja'  # 每次爬之前都要更新的密钥
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E15761066345441398573/ratings'  # 江川右
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E16604909377580131753/ratings'  # 三米1
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E5746667187906760707/ratings'  # 三米2
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E11997536524865771581/ratings'  # 三米3
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E531261138670874370/ratings'  # 三米4
    # url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E2438476065609584660/ratings'  # 三米5
    url = 'https://h5.ele.me/restapi/ugc/v3/restaurants/E3514361176056627367/ratings'  # 三米5

    for i in range(100):
        pizza = get_page(url, i, pizza, sid, save_path)
        time.sleep(3)
